# MediaPrep Studio
# UCF Senior Design Project - Blue Agave
# Application written by Clarisse Vamos for Agri-Starts

#READ ME
#This application was built to run the automated media preparation system designed by Blue Agave of UCF Senior Design.
#It collects user inputs: media type, tray type, media volume in the tank, and the process type. It will run the
#selected process by calling functions from the back-end code. If there are issues with the logic of the system or you
#would like to expand its capabilities, look at the back-end code. Any code edited in this file will affect the front-end
#of the application, or the part of the code that you see running on the screen. This application outputs process results
#to an Excel file in the path directly set in the savedata function. If you need to change where the file is saved, change
#the filepath variable - this will not change the file name.

#Note: the live updating of results was not implemented in this program, but the method create_results is included
#The create_results method when implemented creates a frame on the GUI to show live results
#This method DOES NOT update the application with live results, and that will have to be written if necessary

#Import libraries needed to create application
#import _tkinter
from Tkinter import *
#from Tkinter import ttk
import ttk
#from Tkinter import messagebox
import tkMessageBox
from time import time as tm

#Import libraries needed for saving to excel
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment
from datetime import datetime
import calendar
from pathlib import Path

# Import modules for the backend
from back import Back
from time import sleep
import RPi.GPIO as gpio

# Create Application Root - do not remove
root = Tk()
root.title("MediaPrep Studio")
back = Back()

# Function to ensure processes are stopped when program is closed
def on_closing():
    if tkMessageBox.askokcancel("Quit", "Do you want to quit?"):
        #Make sure the system shuts down
        #class.exit_gracefully()

        print "Closing Application."
        back.motorOff()
        gpio.cleanup()
        root.destroy()

# Set protocol to ensure on main application exit, the process is shut down
root.protocol("WM_DELETE_WINDOW", on_closing)

# Global Variables - these will help keep track of certain values throughout the process
mediaType = "None"          #type of media used in process
trayType = "None"           #type of trays run (jars, tubs, test tubes)
processType = "None"        #type of process carried out
startVol = 0                #volume entered by user as starting volume in tank
numTrays = 0                #number of completed trays
fillVol = 0                 #volume of media that has been used to fill jars
t0=0                        #process start time
rtime = 0                   #process run time
finalt = 0                  #final run time, formatted nicely
errormsg = ""               #error message if needed

# The Application Frame - this class contains the logic for the application
class Application(Frame):

    # Create Initialization Sequence - this initializes the screen the user sees upon open
    def __init__(self, master):
        Frame.__init__(self, master)
        self.create_title()
        self.create_setup()
        #self.create_results() #REMOVED FOR PROTOTYPE
        self.create_buttons()

    # Create Title Frame - this creates the elements for the title
    def create_title(self):
        self.tframe = ttk.Frame(root)
        self.title = ttk.Label(self.tframe, text='Agri-Starts Media Preparation', font = "calibri 16 bold")
        self.title.pack()
        self.tframe.config(padding=(10, 10))
        self.tframe.pack()

    # Create Setup Frame - this creates the elements in the setup frame, where the user will provide inputs
    def create_setup(self):
        #Creates the overall frame that everything will be places in
        self.sframe = ttk.LabelFrame(root, height=200, width=500, text='Setup')
        self.sframe.config(padding=(10, 10))

        #Create the 'Media Type' label
        self.lbl_media = Label(self.sframe, justify=RIGHT, padx=10, text="Media Type:", font="Verdana 10")
        self.lbl_media.grid(column=1,row=1)

        #Media drop down box
        self.media = StringVar(self.sframe)
        self.media.set("None")  # default value
        self.medialist = ttk.Combobox(self.sframe, textvariable=self.media)

        #Change this list to add/subtract media types
        self.medialist['values'] = ('1', '2', '5', '7','10','15','20','30','K1','K3','K5','K7','K10','K1B','K3B','K5B','B1','B5','B10','L','A4','A8','MS','MS 1/2')
        self.medialist.grid(column=2, row=1)

        #Create the 'Media Volume' label
        self.lbl_startVol = Label(self.sframe, justify=RIGHT, padx=10, text="Media Volume (L):", font="Verdana 10")
        self.lbl_startVol.grid(column=1,row=2)

        #Setting the media volume - this is the volume available in the tank
        self.mstartVol = Entry(self.sframe)
        self.mstartVol.grid(column=2,row=2)

        # Create the 'Tray Type' label
        self.lbl_tray = Label(self.sframe, justify=RIGHT, padx=10, text="Tray Type:", font="Verdana 10")
        self.lbl_tray.grid(column=1, row=3)

        # Tray type drop down box - edit the list below to add/remove tray types
        self.trayType = StringVar(self.sframe)
        self.trayType.set("None")  # default value
        self.trayTypelist = OptionMenu(self.sframe, self.trayType, "Jars", "Tubs", "Test Tubes")
        self.trayTypelist.grid(column=2, row=3)

        #Create the 'Process Type' label
        self.lbl_process = Label(self.sframe, justify=RIGHT, padx=10, text="Process Type:", font="Verdana 10")
        self.lbl_process.grid(column=1,row=4)

        #Process drop down box - to add process types, edit this list. Will need to edit elsewhere to implement new process.
        self.process = StringVar(self.sframe)
        self.process.set("None")  # default value
        self.processlist = OptionMenu(self.sframe, self.process, "Automated", "Manual", "Testing")
        self.processlist.grid(column=2, row=4)

        #Pack Frame - place everything within frame.
        self.sframe.pack(fill="both", expand="yes", anchor='w')

    # Create Button Frame - this will create the "Start Process" and "STOP" buttons
    def create_buttons(self):
        #Create the frame to hold the buttons
        self.bframe = ttk.Frame(root)
        self.bframe.config(height=200, width=500)
        self.bframe.config(padding=(10, 10))

        #Create the Process Start Button
        self.select = ttk.Button(self.bframe, text='Start Process', command=self.selectClick)
        self.select.config(state=NORMAL)
        self.select.grid(column=1,row=1)

        #Create the Stop Button
        self.stop = ttk.Button(self.bframe, text = 'STOP', command=self.stopClick)
        self.stop.config(state = DISABLED)
        self.stop.grid(column=2,row=1)

        #Pack Frame
        self.bframe.pack()

    #This method handles the clicking of the 'Start Process' button
    def selectClick(self):
        #retrieve global version variables that will be used
        global t0, mediaType, startVol, processType,trayType, numTrays, fillVol

        #set current time as t0
        t0 = tm()

        #get start values from selections the user made
        mediaType = self.media.get()
        startVol = self.mstartVol.get()
        processType = self.process.get()
        trayType = self.trayType.get()
        fillVol = float(numTrays) * 30 * 30 / 1000

        # this will help check that a media volume is entered
        startL = len(self.mstartVol.get())

        #check for missing information before processing start button click
        if processType == "None" or ((processType == "Automated" and ((startL==0 or startVol==0) or (mediaType== "None" or trayType=="None"))) or (processType == "Manual" and (mediaType == "None" or trayType == "None"))):
            # if there is information missing, create top level pop up window to warn user of missing info
            top = Toplevel(root)
            top.title("Process Start Failed")
            about_message = "Cannot start process due to missing user selection(s). Please correct before attempting to start media preparation."

            # create the window's message
            msg = Message(top, text=about_message, justify=CENTER, pady=5)
            msg.pack()

            # add a button that allows user to exit window
            button = Button(top, text="Dismiss", command=top.destroy, padx=7, pady=5)
            button.pack()

        #No information missing, so change button states and run the process
        else:
            # Change button states
            self.select.config(state=DISABLED)
            self.stop.config(state=NORMAL)

            # Start process
            self.runprocess()

	#This method will run the individual logic of every process
    def runprocess(self):
		# Automated process - run automated line from back-end script
        if (processType == "Automated"):
            print "Running Automated process."
            global fillVol, numTrays
            if gpio.input(4):
                back.motorOff()
                # print "Filling switch triggered"
                sleep(0.25)
                if gpio.input(4):
                    back.motorOn()
                    sleep(3.023)
                    back.motorOff()
                    sleep(1.25)
                    back.filling()
                    sleep(2)

            if gpio.input(27):
                back.motorOff()
                completed += 1
                # print "End switch triggered"
                sleep(0.1)

            else:
                back.motorOn()
                print "running"
                sleep(0.1)

            #Change this to get numtrays from back-end
            numTrays = completed
            fillVol = float(numTrays) * 30 * 30 / 1000

        # Manual Process - run the  line at 30% motor speed from backend script
        elif (processType == "Manual"):
            print "Running Manual process."
            back.motor(18, 300)

        # Testing Process - run test cases from back-end script
        # Only three tests are implemented for the prototype, but more can be added by creating GUI elements and calling back-end test fuctions.
        elif (processType == "Testing"):
            print "Running Testing process."

            # Create the window that will structure the testing mode
            testwindow = Toplevel(root)
            testwindow.title = ("Testing Mode")

            # Add a label, description, and button for each test:
            # Test 1
            lbl_test1 = Label(testwindow, justify=RIGHT, padx=15, pady=5, text="Test 1: Motor Test", font="Verdana 10 bold")
            lbl_test1.grid(column=1, row=1)
            des_test1 = Label(testwindow, justify=RIGHT, padx=15,
                              text="This test will run the motor at different speeds to ensure it is working properly.",
                              font="Verdana 10")
            des_test1.grid(column=1, row=2)
            btn_test1 = Button(testwindow, text="Run Test 1", command=self.test1, padx=10, pady=5)
            btn_test1.grid(column=1, row=3)

            # Test 2
            lbl_test2 = Label(testwindow, justify=RIGHT, padx=15, pady=5, text="Test 2: Limit Switch Test", font="Verdana 10 bold")
            lbl_test2.grid(column=1, row=4)
            des_test2 = Label(testwindow, justify=RIGHT, padx=15,
                              text="This test will run and read inputs to limit switches to ensure they are working properly.",
                              font="Verdana 10")
            des_test2.grid(column=1, row=5)
            btn_test2 = Button(testwindow, text="Run Test 2", command=self.test2, padx=10, pady=5)
            btn_test2.grid(column=1, row=6)

            # Test 3
            lbl_test3 = Label(testwindow, justify=RIGHT, padx=15, pady=5, text="Test 3: Line Movement Test", font="Verdana 10 bold")
            lbl_test3.grid(column=1, row=7)
            des_test3 = Label(testwindow, justify=RIGHT, padx=15,
                              text="This test will run the line without triggering filling to ensure proper positioning and handling of trays.",
                              font="Verdana 10")
            des_test3.grid(column=1, row=8)
            btn_test3 = Button(testwindow, text="Run Test 3", command=self.test3, padx=10, pady=5)
            btn_test3.grid(column=1, row=9)

            # Test 4
            lbl_test4 = Label(testwindow, justify=RIGHT, padx=15, pady=5, text="Test 4: Filling Flush Test",
                              font="Verdana 10 bold")
            lbl_test4.grid(column=1, row=10)
            des_test4 = Label(testwindow, justify=RIGHT, padx=15,
                              text="This test will run a flush of the filling system in order to clean it out.",
                              font="Verdana 10")
            des_test4.grid(column=1, row=11)
            btn_test4 = Button(testwindow, text="Run Test 4", command=self.test4, padx=10, pady=5)
            btn_test4.grid(column=1, row=12)

            #Purely aesthetic
            lbl_pretty = Label(testwindow, justify=RIGHT, padx=15, pady=3, text="", font="Verdana 10 bold")
            lbl_pretty.grid(column=1, row=13)

            # Purely aesthetic
            lbl_remind = Label(testwindow, justify=RIGHT, padx=15, pady=10,
                               text="Note: Remember to still press STOP on main window.", font="Verdana 10")
            lbl_remind.grid(column=1, row=14)

            # Button to finish testing and exit
            btn_closetest = Button(testwindow, text="Finish Test Cases", font="Verdana 10 bold", command=testwindow.destroy, padx=10, pady=5)
            btn_closetest.grid(column=1,row=15)

            print "Running Tests."

        #There is no process type selected - error.
        else:
            print "Error."

    #This method runs test 1 in the testing mode by calling back-end script
    def test1(self):
        #Add method from back-end to run Test 1
        print "Test 1 complete."

    # This method runs test 2 in the testing mode by calling back-end script
    def test2(self):
        # Add method from back-end to run Test 2
        print "Test 2 complete."

    # This method runs test 3 in the testing mode by calling back-end script
    def test3(self):
        # Add method from back-end to run Test 3
        print "Test 3 complete."

    # This method runs test 4 in the testing mode by calling back-end script
    def test4(self):
        # Add method from back-end to run Test 4
        print "Test 4 complete."

    # This handles the logic of pressing the stop button on the application. This will end whatever process
    # in running on the back-end.
    def stopClick(self):

        # Call the back-end exit gracefully function to kill the process currently running
        #class.exit_gracefully()

        # Get the global version of variables and calculate the run time
        global t0, rtime, finalt, errormsg, processType
        rtime = tm() - t0
        finalt = self.formtime()

		#Automated, so just call the save and display method
        if processType == "Automated":
            save = self.save_and_display()
            print "Done with automated process."
            back.motorOff()
            gpio.cleanup()


		#Manual, so take in input on number of trays completed and use that to calculate the filled volume.
        #Then call save and diaplay.
        elif processType == "Manual":
            print "Done with Manual process."
            inputwindow = Toplevel(root)
            inputwindow.title("Input Manual Results")
            inputwindow.geometry("%dx%d%+d%+d" % (300, 200, 250, 125))

            lbl_msg = Label(inputwindow, justify=RIGHT, padx=30, pady=20, text="Input number of trays completed:", font="Verdana 10")
            lbl_msg.pack()

            # Create the 'Trays Completed' label
            lbl_nTrays = Label(inputwindow, justify=RIGHT, padx=15, text="Trays Completed:", font="Verdana 10")
            lbl_nTrays.pack()

            # Set the number of trays completed
            self.trayValue = Entry(inputwindow)
            self.trayValue.pack()

            # Purely aesthetic label
            lbl_pretty = Label(inputwindow, justify=RIGHT, padx=15, text="", font="Verdana 10")
            lbl_pretty.pack()

            submittbtn = Button(inputwindow, text="Done", command=self.save_and_display, padx=10, pady=5)
            submittbtn.pack()

		#Testing, so just state that process is stopped.
        elif processType == "Testing":
            print "Done with Testing process."
            # this should make the stop button grey and start button ready
            resetbtn = Button(finalout, text="Reset Values", command=self.reset, padx=7, pady=5)
            resetbtn.pack()

		#No process entered, so will throw error
        else:
            global errormsg
            errormsg = "No process type."

        # Change button states
        self.stop.config(state=DISABLED)
        self.select.config(state=NORMAL)

	#This method calls the save method and displays the process results to the user in the results pop-up window
    def save_and_display(self):

        # get the global version of variables needed to save and display results
        global mediaType, processType, numTrays, fillVol, finalt, errormsg, trayType, errormsg

        if processType == "Manual":
            numTrays = self.trayValue.get()
        intVol = float(numTrays) * 30 * 30 / 1000
        fillVol = intVol.__str__()

        # save the results
        saveme = self.savedata()

        # Display toplevel window with results
        finalout = Toplevel(root)
        finalout.title("Final Process Results")
        finalout.geometry("%dx%d%+d%+d" % (300, 250, 250, 125))

        # Add a title and message to the window
        about_title = "FINAL PROCESS RESULTS\n\n"
        about_message = "Media Type:\t\t" + mediaType + "\nTray Type:\t\t" + trayType + "\nProcess Type:\t\t" + \
                        processType + "\nTrays Completed:\t\t" + numTrays.__str__() + "\nFilled Volume (L):\t\t" + \
                        fillVol.__str__() + "\nRun Time (H:M:S):\t" + finalt + "\n\n" + errormsg
        if processType == "Manual":
            about_message = about_message + "\nRemember to close all windows that are \nnot the main application window."
        msg = Message(finalout, text=about_title + about_message, width=350, anchor='w', pady=10)
        msg.pack()

        # Close the results pop up and reset all values on screen
        resetbtn = Button(finalout, text="Reset Values", command=self.reset, padx=7, pady=5)
        resetbtn.pack()

	#This method formats the run time to H:M:S
    def formtime(self):
        global rtime
        m, s = divmod(rtime, 60)
        h, m = divmod(m, 60)
        ftime = "%02d:%02d:%02d" % (h, m, s)
        return ftime

	#This method will save the results data to an Excel sheet
    def savedata(self):
        print "Saving...."
        # Get current month and year info
        monthnum = datetime.now().month
        year = datetime.now().year
        month = calendar.month_name[monthnum]
        day = datetime.now().day

        # get global version of variables to save
        global mediaType, processType, numTrays, fillVol, rtime, trayType

        # Create name of data file, corresponding to current year and month. Edit the filepath to change where it saves.
        filename = "MediaPrepData_" + month.__str__() + year.__str__() + ".xlsx"
        filepathend = "\MediaPrepData_" + month.__str__() + year.__str__() + ".xlsx"
        filepath = Path("C:\Users\Clarisse\Documents\MediaPrepStudio" + filepathend)

        try:
            # Open file if it exists in path, otherwise create a new one for the month
            if filepath.is_file():
                # Open workbook for the month
                currdatalog = openpyxl.load_workbook(filename)
                currsheet = currdatalog.get_sheet_by_name("Media Prep Log")

                # Find next available row in file
                rownum = currsheet.max_row + 1

                # Print data to that location
                currsheet.cell(row=rownum, column=1).value = monthnum.__str__() + "/" + day.__str__() + "/" + year.__str__()
                currsheet.cell(row=rownum, column=2).value = mediaType
                currsheet.cell(row=rownum, column=3).value = trayType
                currsheet.cell(row=rownum, column=4).value = processType
                currsheet.cell(row=rownum, column=5).value = numTrays
                currsheet.cell(row=rownum, column=6).value = fillVol
                currsheet.cell(row=rownum, column=7).value = self.formtime()

                # Save the file with updated data
                currdatalog.save(filename)
                print "Save Complete."

            # File not found, so create one
            else:
                print "File not found...creating one."

                # Create workbook for the month
                datalog = Workbook()

                # Create worksheet and give it the proper name
                sheet = datalog.active
                sheet.title = "Media Prep Log"
                print datalog.sheetnames

                # Create heading for file
                sheet.cell(row=1, column=1).value = "Media Prep Log"
                sheet.cell(row=1, column=1).value = "Agri-Starts Media Preparation Data Log"
                sheet.cell(row=2, column=1).value = "Month:"
                sheet.cell(row=2, column=2).value = month.__str__()
                sheet.cell(row=3, column=1).value = "Year:"
                sheet.cell(row=3, column=2).value = year
                sheet.cell(row=3, column=2).alignment = Alignment(horizontal='left')

                # Now create table headers for data
                sheet.cell(row=5, column=1).value = "Date"
                sheet.cell(row=5, column=2).value = "Media Type"
                sheet.cell(row=5, column=3).value = "Tray Type"
                sheet.cell(row=5, column=4).value = "Process Type"
                sheet.cell(row=5, column=5).value = "Trays Completed"
                sheet.cell(row=5, column=6).value = "Filled Volume (L)"
                sheet.cell(row=5, column=7).value = "Run Time (H:M:S)"

                # Set column widths
                sheet.column_dimensions['A'].width = 12.0
                sheet.column_dimensions['B'].width = 12.0
                sheet.column_dimensions['C'].width = 13.0
                sheet.column_dimensions['D'].width = 13.0
                sheet.column_dimensions['E'].width = 13.0
                sheet.column_dimensions['F'].width = 16.0
                sheet.column_dimensions['G'].width = 15.0
                sheet.column_dimensions['H'].width = 16.0

                # Set row number to 6, which is start of data
                rownum=6

                # Print data to that location
                sheet.cell(row=rownum, column=1).value = monthnum.__str__() + "/" + day.__str__() + "/" + year.__str__()
                sheet.cell(row=rownum, column=2).value = mediaType
                sheet.cell(row=rownum, column=3).value = trayType
                sheet.cell(row=rownum, column=4).value = processType
                sheet.cell(row=rownum, column=5).value = numTrays
                sheet.cell(row=rownum, column=6).value = fillVol
                sheet.cell(row=rownum, column=7).value = self.formtime()

                # Save file
                datalog.save(filename)
                print "Save Complete."

        #There is an issue saving the file, so set message that will prompt the user to manually save data.
        except Exception:
            global errormsg
            errormsg = "Problem saving to file. Please manually enter data \ninto file if you would like it to be logged."
            return errormsg
        return ""

	# This performs a reset of all values on the main application window
    # It ensures that data is not kept from one process to the next
    def reset(self):
        print "Values are reset."
        # Get the global variables and reset them
        global mediaType, trayType, processType, startVol, numTrays, fillVol, t0, rtime, finalt, errormsg
        mediaType = "None"
        trayType = "None"
        processType = "None"
        startVol = 0
        numTrays = 0
        fillVol = 0
        t0 = 0
        rtime = 0
        finalt = 0
        errormsg = ""

        # Reset the input values
        self.media.set("None")
        mediaType = "None"
        self.mstartVol.delete(0, 'end')
        self.trayValue.delete(0,'end')
        self.process.set("None")
        self.trayType.set("None")

        # Change button states
        self.stop.config(state=DISABLED)
        self.select.config(state=NORMAL)

    # Create Current Process Frame - REMOVED FOR PROTOTYPE
    # If implemted, the grid col, row for the button frame will have to be edited
    def create_results(self):
        self.rframe = ttk.LabelFrame(root, height=200, width=500, text='Current Process')
        self.rframe.config(padding=(10, 10))

        # Create the 'Media Type' label
        self.lbl_media2 = Label(self.rframe, justify=RIGHT, padx=10, text="Media Type:", font="Verdana 10")
        self.lbl_media2.grid(column=1, row=1)

        # Create Media Type value label
        self.mediaValue = Label(self.rframe, justify=LEFT, padx=10, text="")
        self.mediaValue.grid(column=2, row=1)

        # Create the 'Process Type' label
        self.lbl_process2 = Label(self.rframe, justify=RIGHT, padx=10, text="Process Type:", font="Verdana 10")
        self.lbl_process2.grid(column=1, row=2)

        # Create Process Type value label
        self.processValue = Label(self.rframe, justify=LEFT, padx=10, text="")
        self.processValue.grid(column=2, row=2)

        # Create the 'Tray Type' label
        self.lbl_trayType = Label(self.rframe, justify=RIGHT, padx=10, text="Tray Type:", font="Verdana 10")
        self.lbl_trayType.grid(column=1, row=3)

        # Create Trays Completed value label
        self.trayTypeValue = Label(self.rframe, justify=LEFT, padx=10, text="")
        self.trayTypeValue.grid(column=2, row=3)

        # Create the 'Trays Completed' label
        self.lbl_numTrays = Label(self.rframe, justify=RIGHT, padx=10, text="Trays Completed:", font="Verdana 10")
        self.lbl_numTrays.grid(column=1, row=4)

        # Create Trays Completed value label
        self.trayValue = Label(self.rframe, justify=LEFT, padx=10, text="")
        self.trayValue.grid(column=2, row=4)

        # Create the 'Filled Volume' label
        self.lbl_fillVol = Label(self.rframe, justify=RIGHT, padx=10, text="Filled Volume (L):", font="Verdana 10")
        self.lbl_fillVol.grid(column=1, row=5)

        # Create Filled Volume value label
        self.fillValue = Label(self.rframe, justify=LEFT, padx=10, text="")
        self.fillValue.grid(column=2, row=5)

        # Pack Frame
        self.rframe.pack(fill="both", expand="yes", anchor='w')

# End of application - DO NOT EDIT
app = Application(root)
root.mainloop()
